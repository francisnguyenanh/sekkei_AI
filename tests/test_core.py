# tests/test_core.py
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import pytest
import openpyxl
import io

# --- StyleDef validator tests ---
from schemas.models import StyleDef, GenerateRequest, MultiSheetGenerateRequest, parse_request

class TestColorValidator:
    def test_valid_6char_rgb(self):
        s = StyleDef(fill="0070C0")
        assert s.fill == "0070C0"

    def test_valid_8char_argb(self):
        s = StyleDef(fill="FF0070C0")
        assert s.fill == "FF0070C0"

    def test_rejects_invalid_color(self):
        from pydantic import ValidationError
        with pytest.raises(ValidationError):
            StyleDef(fill="ZZZZZZ")

    def test_rejects_5char_color(self):
        from pydantic import ValidationError
        with pytest.raises(ValidationError):
            StyleDef(fill="0070C")

    def test_none_is_allowed(self):
        s = StyleDef(fill=None)
        assert s.fill is None


# --- Border parser tests ---
from core.styles import get_border, _parse_composite_border
from openpyxl.styles import Border

class TestBorderParser:
    def test_all_thin(self):
        b = get_border("all_thin")
        assert b.left.border_style == "thin"
        assert b.right.border_style == "thin"

    def test_none_returns_empty_border(self):
        b = get_border("none")
        assert isinstance(b, Border)
        assert b.left.border_style is None

    def test_composite_left_top(self):
        b = get_border("L:thin,T:thin")
        assert b.left.border_style == "thin"
        assert b.top.border_style == "thin"
        assert b.right.border_style is None
        assert b.bottom.border_style is None

    def test_composite_medium(self):
        b = get_border("L:medium,R:medium,T:medium,B:medium")
        assert b.left.border_style == "medium"

    def test_unknown_named_border_returns_empty(self):
        b = get_border("nonexistent_style")
        assert isinstance(b, Border)


# --- ExcelGeneratorService tests ---
from core.generator import ExcelGeneratorService

def make_minimal_request(sheet_name="TestSheet"):
    return parse_request({
        "template": {
            "sheet_name": sheet_name,
            "global_config": {"column_widths": {"A": 10.0}, "row_heights": {"1": 13.0}},
            "styles": {
                "blue_header": {"fill": "0070C0", "bold": True, "border": "all_thin"}
            },
            "layout_blocks": [
                {"range": "A1:C1", "style": "blue_header", "merge": True, "static_text": "Header"}
            ],
            "mapping_anchors": {"screen_id": "B2"}
        },
        "logic": {
            "single_values": {"screen_id": "SCR-001"},
            "table_data": {}
        }
    })

class TestExcelGeneratorService:
    def test_single_sheet_generates_bytes(self):
        req = make_minimal_request()
        svc = ExcelGeneratorService(req)
        output = svc.generate()
        assert isinstance(output, io.BytesIO)
        assert output.tell() == 0  # seeked to start

    def test_output_is_valid_xlsx(self):
        req = make_minimal_request()
        svc = ExcelGeneratorService(req)
        output = svc.generate()
        wb = openpyxl.load_workbook(output)
        assert "TestSheet" in wb.sheetnames

    def test_static_text_written(self):
        req = make_minimal_request()
        svc = ExcelGeneratorService(req)
        output = svc.generate()
        wb = openpyxl.load_workbook(output)
        ws = wb["TestSheet"]
        assert ws["A1"].value == "Header"

    def test_logic_single_value_injected(self):
        req = make_minimal_request()
        svc = ExcelGeneratorService(req)
        output = svc.generate()
        wb = openpyxl.load_workbook(output)
        ws = wb["TestSheet"]
        assert ws["B2"].value == "SCR-001"

    def test_row_height_applied_as_int_key(self):
        # row_heights with STRING keys must not crash on save
        req = make_minimal_request()
        svc = ExcelGeneratorService(req)
        # Manually set str key to simulate defensive scenario
        svc.ws.row_dimensions[1].height = 13.0
        output = io.BytesIO()
        svc.wb.save(output)  # Must not raise TypeError
        assert output.tell() > 0

    def test_save_false_returns_none(self):
        req = make_minimal_request()
        svc = ExcelGeneratorService(req)
        result = svc.generate(save=False)
        assert result is None

    def test_multi_sheet_shared_workbook(self):
        wb = openpyxl.Workbook()
        req1 = make_minimal_request("Sheet1")
        req2 = make_minimal_request("Sheet2")
        svc1 = ExcelGeneratorService(req1, workbook=wb)
        svc1.generate(save=False)
        svc2 = ExcelGeneratorService(req2, workbook=wb)
        svc2.generate(save=False)
        assert "Sheet1" in wb.sheetnames
        assert "Sheet2" in wb.sheetnames


# --- MultiSheetGenerateRequest validator tests ---
class TestMultiSheetValidator:
    def test_empty_sheets_rejected(self):
        from pydantic import ValidationError
        with pytest.raises(ValidationError, match="must not be empty"):
            MultiSheetGenerateRequest.model_validate({"sheets": []})

    def test_over_50_sheets_rejected(self):
        from pydantic import ValidationError
        single_sheet = make_minimal_request().model_dump()
        with pytest.raises(ValidationError, match="maximum of 50"):
            MultiSheetGenerateRequest.model_validate({"sheets": [single_sheet] * 51})

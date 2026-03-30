import json
import sys
import argparse
import openpyxl
from openpyxl.utils.cell import get_column_letter

def main():
    parser = argparse.ArgumentParser(description="Extract JSON Template from Excel file")
    parser.add_argument("input", help="Path to input .xlsx file")
    parser.add_argument("output", help="Path to save output .json file")
    parser.add_argument("--sheet", help="Name of the sheet to extract", default=None)
    parser.add_argument("--data-start-row", type=int, default=49, help="Row number where table data starts (styles below this are ignored)")
    parser.add_argument("--anchors", help="Path to a JSON file containing mapping_anchors dict", default=None)
    
    args = parser.parse_args()
    
    sys.stderr.write(f"Loading '{args.input}'...\n")
    try:
        wb = openpyxl.load_workbook(args.input, data_only=True)
    except Exception as e:
        sys.stderr.write(f"Error loading Excel: {e}\n")
        sys.exit(1)
        
    sheet_name = args.sheet if args.sheet else wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        sys.stderr.write(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}\n")
        sys.exit(1)
        
    ws = wb[sheet_name]
    sys.stderr.write(f"Extracting sheet '{sheet_name}', ignoring styles > row {args.data_start_row}...\n")

    styles_dict = {}
    layout_blocks = []
    
    col_widths = {}
    max_col = ws.max_column or 0
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        cd = ws.column_dimensions.get(letter)
        if cd and cd.width:
            col_widths[letter] = cd.width

    row_heights = {}
    max_row = ws.max_row or 0
    for row_num in range(1, max_row + 1):
        h = ws.row_dimensions[row_num].height
        if h is not None:
            row_heights[row_num] = h

    if args.anchors:
        with open(args.anchors, encoding='utf-8') as f:
            mapping_anchors = json.load(f)
    else:
        mapping_anchors = {}
        sys.stderr.write("Warning: mapping_anchors is empty. Pass --anchors <file.json> to include anchor definitions.\n")

    def get_border_key(cell) -> str | None:
        b = cell.border
        if not b:
            return None
        side_map = {"thin": "thin", "medium": "medium", None: "none"}
        parts = []
        mapping = [("L", b.left), ("R", b.right), ("T", b.top), ("B", b.bottom)]
        has_any = False
        for key, side in mapping:
            style = side.border_style if side else None
            if style in ("thin", "medium"):
                parts.append(f"{key}:{style}")
                has_any = True
        return ",".join(parts) if has_any else None

    def get_style_hash(cell):
        fill = cell.fill.fgColor.rgb if (cell.fill and type(cell.fill.fgColor).__name__ == 'Color' and type(cell.fill.fgColor.rgb) == str) else None
        if fill == '00000000': fill = None
        
        font_color = cell.font.color.rgb if (cell.font and hasattr(cell.font, 'color') and cell.font.color and type(cell.font.color.rgb) == str) else None
        if font_color == '00000000': font_color = None

        font_size = cell.font.sz if cell.font else None
        bold = cell.font.bold if cell.font else False
        font_name = cell.font.name if cell.font else None
        
        align_h = cell.alignment.horizontal if cell.alignment else None
        align_v = cell.alignment.vertical if cell.alignment else None
        
        wrap_text = cell.alignment.wrap_text if cell.alignment else None
        if wrap_text is False:
            wrap_text = None
            
        border = get_border_key(cell)

        if not any([fill, font_color, bold, align_h, align_v, wrap_text, border]) and (font_size == 9.0 or font_size is None):
            return None

        style_cfg = {
            "fill": fill,
            "font_color": font_color,
            "font_size": font_size,
            "font_name": font_name,
            "bold": bold,
            "align_h": align_h,
            "align_v": align_v,
            "wrap_text": wrap_text,
            "border": border
        }
        
        return style_cfg, str(style_cfg)

    style_counter = 1
    hash_to_name = {}
    processed_cells = set()

    for range_str in [str(r) for r in ws.merged_cells.ranges]:
        tl_coord = range_str.split(":")[0]
        cell = ws[tl_coord]
        
        for row in ws[range_str]:
            for c in row:
                processed_cells.add(c.coordinate)
                
        if cell.row >= args.data_start_row:
            continue

        style_tuple = get_style_hash(cell)
        style_name = None
        if style_tuple:
            style_cfg, shash = style_tuple
            if shash not in hash_to_name:
                style_name = f"style_{style_counter:03d}"
                styles_dict[style_name] = style_cfg
                hash_to_name[shash] = style_name
                style_counter += 1
            else:
                style_name = hash_to_name[shash]
        
        static_text = None
        if cell.value is not None:
            static_text = str(cell.value)
            
        if style_name or static_text:
            layout_blocks.append({
                "range": range_str,
                "style": style_name,
                "merge": True,
                "static_text": static_text
            })

    for row in ws.iter_rows(min_row=1, max_row=args.data_start_row - 1):
        for cell in row:
            if cell.coordinate in processed_cells:
                continue
                
            style_tuple = get_style_hash(cell)
            static_text = None
            if cell.value is not None:
                static_text = str(cell.value)
                
            style_name = None
            if style_tuple:
                style_cfg, shash = style_tuple
                if shash not in hash_to_name:
                    style_name = f"style_{style_counter:03d}"
                    styles_dict[style_name] = style_cfg
                    hash_to_name[shash] = style_name
                    style_counter += 1
                else:
                    style_name = hash_to_name[shash]
                    
            if style_name or static_text:
                layout_blocks.append({
                    "range": cell.coordinate,
                    "style": style_name,
                    "merge": False,
                    "static_text": static_text
                })

    out = {
        "sheet_name": sheet_name,
        "global_config": {
            "column_widths": col_widths,
            "row_heights": row_heights,
            "default_font": {"name": "Meiryo UI", "size": 9.0}
        },
        "styles": styles_dict,
        "layout_blocks": layout_blocks,
        "mapping_anchors": mapping_anchors
    }

    sys.stderr.write(f"Saving to {args.output}...\n")
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
        
    sys.stderr.write("Done!\n")

if __name__ == "__main__":
    main()

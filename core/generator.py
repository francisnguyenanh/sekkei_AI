import io
import copy
import logging
import openpyxl
from openpyxl.utils.cell import get_column_letter, coordinate_from_string, column_index_from_string
from core.styles import create_named_style
from schemas.models import GenerateRequest

logger = logging.getLogger(__name__)

class ExcelGeneratorService:
    def __init__(self, request: GenerateRequest, workbook=None):
        self.template = request.template
        self.logic = request.logic
        self.wb = workbook if workbook else openpyxl.Workbook()
        
        if workbook is None:
            # Fresh workbook: rename the default sheet
            self.ws = self.wb.active
            self.ws.title = self.template.sheet_name
        else:
            # Shared workbook: remove the default empty sheet if it's still there and untouched
            default_sheets = [s for s in self.wb.sheetnames if s in ("Sheet", "Sheet1")]
            if default_sheets and len(self.wb.sheetnames) == 1:
                self.ws = self.wb.active
                self.ws.title = self.template.sheet_name
            else:
                self.ws = self.wb.create_sheet(title=self.template.sheet_name)
            
        self.precompiled_styles = {}

    def _compile_styles(self):
        for name, style_def in self.template.styles.items():
            self.precompiled_styles[name] = create_named_style(name, style_def)

    def _apply_global_config(self):
        # Set column widths
        for col_name, width in self.template.global_config.column_widths.items():
            self.ws.column_dimensions[col_name].width = width
            
        # Set row heights
        for row_num, height in self.template.global_config.row_heights.items():
            self.ws.row_dimensions[int(row_num)].height = height

    def _apply_layout_blocks(self):
        for block in self.template.layout_blocks:
            cells = self.ws[block.range]
            
            # Normalize cells to tuple of tuples for iteration
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            elif not isinstance(cells[0], tuple):
                cells = (cells,)
                
            style_kwargs = self.precompiled_styles.get(block.style, {})

            for row in cells:
                for cell in row:
                    if 'fill' in style_kwargs: cell.fill = copy.copy(style_kwargs['fill'])
                    if 'font' in style_kwargs: cell.font = copy.copy(style_kwargs['font'])
                    if 'alignment' in style_kwargs: cell.alignment = copy.copy(style_kwargs['alignment'])
                    if 'border' in style_kwargs: cell.border = copy.copy(style_kwargs['border'])
                    
                    if block.static_text and cell.coordinate == block.range.split(":")[0]:
                        cell.value = block.static_text

            if block.merge and ":" in block.range:
                self.ws.merge_cells(block.range)

    def _inject_logic_data(self):
        # Validate missing logic anchors
        missing = set(self.logic.single_values.keys()) - set(self.template.mapping_anchors.keys())
        missing_table = set(self.logic.table_data.keys()) - set(self.template.mapping_anchors.keys())
        if missing or missing_table:
            logger.warning("These logic keys have no anchor in template: %s", missing | missing_table)

        # Inject single values
        for key, value in self.logic.single_values.items():
            if key in self.template.mapping_anchors:
                coord = self.template.mapping_anchors[key]
                try:
                    self.ws[coord] = value
                except AttributeError:
                    logger.warning("Skipped merged cell at %s key=%s", coord, key)

        # Inject table data
        for key, rows in self.logic.table_data.items():
            if key in self.template.mapping_anchors:
                start_coord = self.template.mapping_anchors[key]
                xy = coordinate_from_string(start_coord)
                start_col = column_index_from_string(xy[0])
                start_row = xy[1]

                for row_idx, row_data in enumerate(rows):
                    for col_idx, cell_value in enumerate(row_data):
                        try:
                            self.ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=cell_value)
                        except AttributeError:
                            logger.warning(
                                "Skipped merged cell at row=%d col=%d key=%s",
                                start_row + row_idx, start_col + col_idx, key
                            )

    def generate(self, save: bool = True):
        self._compile_styles()
        self._apply_global_config()
        self._apply_layout_blocks()
        self._inject_logic_data()

        if not save:
            return None

        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
